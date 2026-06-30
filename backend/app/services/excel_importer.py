"""Import historical data from Excel file (ĐẶT CƠM 2026.xlsx)."""

import re
from datetime import date, datetime, timezone
from typing import IO

import openpyxl

COL_MEMBER_NAME = 2
COL_DISH = 3
COL_DISH_ALT = 4  # formerly "chay" column, used as fallback
COL_NOTE = 6
COL_EXTRA = 9

DATE_PATTERN = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")


def _parse_sheet_date(sheet_name: str) -> date | None:
    m = DATE_PATTERN.match(sheet_name.strip())
    if not m:
        return None
    try:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def _parse_extra(cell_value) -> tuple[str | None, int]:
    if not cell_value:
        return None, 0
    val = str(cell_value).strip()
    if not val:
        return None, 0
    for sep in ["=", ":"]:
        if sep in val:
            parts = val.split(sep, 1)
            desc = parts[0].strip()
            try:
                cost = int(float(parts[1].strip().replace("k", "").replace(",", ".")))
                return desc, cost
            except (ValueError, IndexError):
                pass
    return val, 0


def _parse_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(",", ".").replace("k", "")))
    except (ValueError, TypeError):
        return 0


def preview_import(file: IO[bytes], member_name_map: dict[str, int] | None = None) -> dict:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    days = []
    unknown_names = set()
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        order_date = _parse_sheet_date(sheet_name)
        if not order_date:
            skipped_sheets.append(sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        total_bill = 0
        for row in rows[:5]:
            for cell in row:
                v = _parse_int(cell)
                if 50 <= v <= 5000:
                    total_bill = v
                    break
            if total_bill:
                break

        members_in_day = []
        for row in rows[2:]:
            name_val = row[COL_MEMBER_NAME - 1] if len(row) >= COL_MEMBER_NAME else None
            if not name_val:
                continue
            name = str(name_val).strip()
            if not name or name.lower() in ("tên", "name", "stt", ""):
                continue

            dish = str(row[COL_DISH - 1]).strip() if len(row) >= COL_DISH and row[COL_DISH - 1] else ""
            dish_alt = str(row[COL_DISH_ALT - 1]).strip() if len(row) >= COL_DISH_ALT and row[COL_DISH_ALT - 1] else ""
            note = str(row[COL_NOTE - 1]).strip() if len(row) >= COL_NOTE and row[COL_NOTE - 1] else ""
            extra_val = row[COL_EXTRA - 1] if len(row) >= COL_EXTRA else None
            extra_desc, extra_cost = _parse_extra(extra_val)

            dish_name = dish or dish_alt
            is_eating = bool(dish_name)

            member_id = (member_name_map or {}).get(name)
            if is_eating and not member_id:
                unknown_names.add(name)

            members_in_day.append({
                "excel_name": name,
                "member_id": member_id,
                "dish": dish_name,
                "note": note or None,
                "extra_desc": extra_desc,
                "extra_cost": extra_cost,
                "is_eating": is_eating,
            })

        days.append({"date": str(order_date), "total_bill": total_bill, "members": members_in_day})

    days.sort(key=lambda d: d["date"])
    return {
        "days": days,
        "unknown_names": sorted(unknown_names),
        "skipped_sheets": skipped_sheets,
        "total_days": len(days),
    }


def commit_import(file: IO[bytes], member_name_map: dict[str, int], db) -> dict:
    """Import data into Firestore. Skips days that already exist."""
    from app.database import _next_id

    preview = preview_import(file, member_name_map)
    imported = 0
    skipped = 0
    errors = []

    for day in preview["days"]:
        order_date = day["date"]

        # Check if this date already exists
        if db.collection("daily_orders").document(order_date).get().exists:
            skipped += 1
            continue

        try:
            now = datetime.now(timezone.utc)
            order_id = _next_id(db, "daily_orders")
            db.collection("daily_orders").document(order_date).set({
                "id": order_id,
                "order_date": order_date,
                "status": "finalized",
                "total_bill": day["total_bill"],
                "shared_cost_per_person": 0,
                "order_deadline": None,
                "note": None,
                "created_by": None,
                "created_at": now,
            })

            batch = db.batch()
            for m in day["members"]:
                mid = m.get("member_id")
                if not mid:
                    continue
                item_id = _next_id(db, "order_items")
                ref = db.collection("order_items").document(str(item_id))
                batch.set(ref, {
                    "id": item_id,
                    "daily_order_id": order_id,
                    "order_date": order_date,
                    "member_id": mid,
                    "dish_name": m["dish"],
                    "note": m["note"],
                    "is_eating": m["is_eating"],
                    "extra_item_description": m["extra_desc"],
                    "extra_item_cost": m["extra_cost"],
                    "total_cost": 0,
                    "created_at": now,
                })
            batch.commit()
            imported += 1
        except Exception as e:
            errors.append(f"{order_date}: {str(e)}")

    return {"imported": imported, "skipped": skipped, "errors": errors}
